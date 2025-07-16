#!/usr/bin/python3
import subprocess
import json
import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import calendar
import logging
import time # Importar o módulo time

# Configuração do Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Datas de consulta (Dinâmicas)
today = datetime.now()
inicio_ano = datetime(today.year, 1, 1).strftime("%Y-%m-%d") # Início do ano corrente

# Datas fixas para Junho
junho_do_ano_atual = datetime(today.year, 6, 1)
inicio_mes_atual = junho_do_ano_atual.strftime("%Y-%m-%d") # Início de Junho do ano corrente
fim_mes_atual = datetime(today.year, 6, 30).strftime("%Y-%m-%d") # Fim de Junho do ano corrente
meses_passados_no_ano = 6 # Fixado para 6 meses (Jan a Jun)

fim_ano = datetime(today.year, 12, 31).strftime("%Y-%m-%d") # Fim do ano corrente

logging.info(f"Gerando relatório de custos para o período: {inicio_mes_atual} a {fim_mes_atual} (Gasto Atual).")
logging.info(f"Previsão até dezembro/{today.year} baseada na média de gastos de {inicio_ano} a {fim_mes_atual}.")

# Cria planilha
wb = Workbook()
ws = wb.active
ws.title = "Custos Azure"

# Cabeçalhos (adaptados para as datas dinâmicas)
headers = ["Subscription Name", "Subscription ID", f"Gasto Total YTD até Junho/{today.year}", f"Previsão até dezembro/{today.year}"]
ws.append(headers)

# Estilo do cabeçalho
header_fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
for col in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center")

# Obtem todas as subscriptions via Azure CLI
try:
    subs_raw = subprocess.check_output(["az", "account", "list", "--query", "[].{name:name, id:id, tenantId:tenantId}", "-o", "json"])
    subscriptions = json.loads(subs_raw)
    logging.info(f"Total de {len(subscriptions)} subscriptions Azure encontradas.")
except subprocess.CalledProcessError as e:
    logging.error(f"Erro ao obter subscriptions do Azure CLI. Certifique-se de estar logado: {e}")
    subscriptions = []
except json.JSONDecodeError as e:
    logging.error(f"Erro ao parsear JSON das subscriptions do Azure CLI: {e}")
    subscriptions = []

# Corpo base para consulta da API de Custo (apenas para 'ActualCost')
def build_cost_query(time_from, time_to):
    return {
        "type": "ActualCost", # Sempre "ActualCost" para este endpoint
        "timeframe": "Custom",
        "timePeriod": {
            "from": time_from,
            "to": time_to
        },
        "dataset": {
            "granularity": "None",
            "aggregation": {
                "totalCost": {
                    "name": "PreTaxCost",
                    "function": "Sum"
                }
            }
        }
    }

# Configurações de retry para HTTP 429
MAX_RETRIES = 3
RETRY_DELAY_SECONDS = 5 # Atraso inicial entre retries

# Faz chamadas para cada subscription
for sub in subscriptions:
    name = sub["name"]
    sub_id = sub["id"]
    tenant_id = sub["tenantId"]
    
    actual_cost_url = f"https://management.azure.com/subscriptions/{sub_id}/providers/Microsoft.CostManagement/query?api-version=2021-10-01"

    gasto_ytd = "N/A"
    previsao_final_ano = "N/A"
    
    logging.info(f"Processando subscription: {name} (ID: {sub_id}, Tenant: {tenant_id})")

    token = None
    headers_api = {}
    try:
        subprocess.check_output(["az", "account", "set", "--subscription", sub_id])
        logging.info(f"Contexto do Azure CLI definido para subscription: {sub_id}")

        token_raw = subprocess.check_output(["az", "account", "get-access-token", "--resource", "https://management.azure.com", "--tenant", tenant_id, "-o", "json"])
        token_info = json.loads(token_raw)
        token = token_info["accessToken"]
        
        expires_on_str = token_info["expiresOn"]
        try:
            expires_on = datetime.strptime(expires_on_str, "%Y-%m-%d %H:%M:%S.%f")
        except ValueError:
            try:
                expires_on = datetime.strptime(expires_on_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                logging.warning(f"Não foi possível parsear a string de expiração do token para {name}: '{expires_on_str}'. Usando a string bruta.")
                expires_on = expires_on_str

        logging.info(f"Token para {name} obtido com sucesso. Válido até: {expires_on}")
        headers_api = {
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        }
    except subprocess.CalledProcessError as e:
        error_output = e.stderr.decode().strip() if e.stderr else str(e)
        logging.error(f"Erro ao definir o contexto ou obter token para {name} ({sub_id}): {error_output}")
        gasto_ytd = "Erro Auth"
        previsao_final_ano = "Erro Auth"
        ws.append([name, sub_id, gasto_ytd, previsao_final_ano])
        continue 
    except json.JSONDecodeError as e:
        logging.error(f"Erro ao parsear JSON do token para {name} ({sub_id}): {e}")
        gasto_ytd = "Erro Parse Token"
        previsao_final_ano = "Erro Parse Token"
        ws.append([name, sub_id, gasto_ytd, previsao_final_ano])
        continue
    except Exception as e:
        logging.critical(f"Erro inesperado no processo de autenticação para {name} ({sub_id}): {e}")
        gasto_ytd = "Erro Inesperado Auth"
        previsao_final_ano = "Erro Inesperado Auth"
        ws.append([name, sub_id, gasto_ytd, previsao_final_ano])
        continue

    # --- INÍCIO DA LÓGICA DE RETRY PARA O GASTO YTD ---
    for attempt in range(MAX_RETRIES):
        try:
            logging.info(f"Consultando gasto YTD para: {name} ({sub_id}) de {inicio_ano} a {fim_mes_atual} (Tentativa {attempt + 1}/{MAX_RETRIES})")
            actual_req = requests.post(actual_cost_url, headers=headers_api, json=build_cost_query(inicio_ano, fim_mes_atual), timeout=60) # Aumentei o timeout
            
            if actual_req.status_code == 200:
                actual_data = actual_req.json()
                if actual_data.get("properties") and actual_data["properties"].get("rows"):
                    gasto_ytd_valor = actual_data["properties"]["rows"][0][0]
                    gasto_ytd = gasto_ytd_valor
                    logging.info(f"Gasto YTD para {name} ({sub_id}): {gasto_ytd}")

                    if meses_passados_no_ano > 0 and isinstance(gasto_ytd_valor, (int, float)):
                        custo_medio_mensal = gasto_ytd_valor / meses_passados_no_ano
                        meses_restantes = 12 - meses_passados_no_ano
                        previsao_futura = custo_medio_mensal * meses_restantes
                        previsao_final_ano = gasto_ytd_valor + previsao_futura
                        logging.info(f"Previsão por média para {name} ({sub_id}): {previsao_final_ano}")
                    else:
                        previsao_final_ano = "Sem Dados para Prever"
                        logging.info(f"Não há meses suficientes ou dados numéricos para calcular a previsão por média para {name}.")
                else:
                    logging.info(f"Nenhum dado de gasto YTD encontrado para {name} ({sub_id}) de {inicio_ano} a {fim_mes_atual}. Resposta: {actual_data}")
                    gasto_ytd = "Sem Dados"
                    previsao_final_ano = "Sem Dados"
                break # Sai do loop de retry se a requisição for bem-sucedida
            elif actual_req.status_code == 429:
                logging.warning(f"Recebido HTTP 429 (Too Many Requests) para {name} ({sub_id}). Tentando novamente em {RETRY_DELAY_SECONDS * (attempt + 1)} segundos...")
                time.sleep(RETRY_DELAY_SECONDS * (attempt + 1)) # Aumenta o delay a cada tentativa
            else:
                logging.error(f"Falha ao obter gasto YTD para {name} ({sub_id}). Status: {actual_req.status_code}, Resposta: {actual_req.text}")
                gasto_ytd = f"Erro HTTP {actual_req.status_code}"
                previsao_final_ano = f"Erro HTTP {actual_req.status_code}"
                break # Sai do loop de retry para outros erros HTTP
        except requests.exceptions.RequestException as e:
            logging.error(f"Erro de rede/conexão ao consultar gasto YTD para {name} ({sub_id}): {e}")
            gasto_ytd = "Erro API"
            previsao_final_ano = "Erro API"
            if attempt == MAX_RETRIES - 1: # Se for a última tentativa
                break # Sai do loop de retry
            else:
                time.sleep(RETRY_DELAY_SECONDS * (attempt + 1)) # Espera para tentar novamente
        except (KeyError, IndexError) as e:
            logging.error(f"Erro ao parsear dados JSON de gasto YTD para {name} ({sub_id}): {e}. Dados recebidos: {actual_data.get('properties', 'N/A') if 'actual_data' in locals() else 'No data received'}")
            gasto_ytd = "Erro Parse"
            previsao_final_ano = "Erro Parse"
            break # Sai do loop de retry para erro de parsing
        except Exception as e:
            logging.critical(f"Erro inesperado ao processar gasto YTD para {name} ({sub_id}): {e}")
            gasto_ytd = "Erro Inesperado"
            previsao_final_ano = "Erro Inesperado"
            break # Sai do loop de retry para erros inesperados
    # --- FIM DA LÓGICA DE RETRY ---

    # Adiciona linha na planilha
    ws.append([
        name,
        sub_id,
        gasto_ytd,
        previsao_final_ano
    ])

# Formatação das colunas numéricas
for row in ws.iter_rows(min_row=2, min_col=3, max_col=4):
    for cell in row:
        if isinstance(cell.value, (float, int)):
            cell.number_format = 'R$ #,##0.00'
            cell.alignment = Alignment(horizontal="right")
        else:
            cell.alignment = Alignment(horizontal="center")

# Auto-ajusta largura das colunas
for column_cells in ws.columns:
    max_length = 0
    column = column_cells[0].column_letter
    for cell in column_cells:
        try:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        except TypeError:
            max_length = max(max_length, len(str(cell.value)))
    adjusted_width = (max_length + 2)
    if column == 'C':
        adjusted_width = max(adjusted_width, len(headers[2]) + 2)
    elif column == 'D':
        adjusted_width = max(adjusted_width, len(headers[3]) + 2)

    ws.column_dimensions[column].width = adjusted_width

# Salva o arquivo
arquivo = f"custos_azure_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
try:
    wb.save(arquivo)
    logging.info(f"Planilha de custos Azure salva com sucesso como: {arquivo}")
    print(f"Planilha de custos Azure salva como: {arquivo}")
except Exception as e:
    logging.critical(f"Erro ao salvar a planilha: {e}")
    print(f"ERRO: Não foi possível salvar a planilha. Detalhes: {e}")